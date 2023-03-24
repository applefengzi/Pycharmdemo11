# -*- coding: UTF-8 -*-
from time import time, sleep
from turtle import back
import requests
import json
from openpyxl import load_workbook
import unittest


class Statistics():
    def sacve(self,data_list):
        excel_path = "统计验证.xlsx"  #表格名称

        # wb = Workbook(excel_path)             #创建xlsx表格
        # wb.save(excel_path)                   #保存

        wb = load_workbook(excel_path)        #导入表格，用于下面进行操作
        wb.create_sheet("验证结果") #创建表名
        ws = wb["验证结果"]                       #激活
        data_list = [['功能介绍', 'opa', 'record', 'id', None, None],['功能介绍1', 'opa1', 'record1', 'id1', None, None]]
        print(len(data_list))
        for row in range(1,len(data_list)+1):   #控制行
            print(row)
            for column in range(1,6):         #控制列
                print(column)
                print(data_list[row-1][column-1])
                ws.cell(row,column).value = data_list[row-1][column-1]  #把数据传输到表格中

        wb.save(excel_path)                   #保存表格
        wb.close()                            #关闭

        print("%s     保存成功！" % excel_path)


    def dowm(self):
        self.las = []

        wb = load_workbook("D:\pack\统计验证.xlsx")
        sheet = wb.active  # 获取活动表
        print('=================================================')
        # las = []
        for row in sheet:  # 循环获取表数据
            lls = []
            for cell in row:  # 循环获取每个单元格数据
                lls.append(cell.value)
                # print(cell.value, end=",")
                # print(lls)
            self.las.append(lls)
            # print()
        # print(self.las)

        self.a = self.get_json()

    def get_json(self):
        url = 'http://hdhlogs.datacenter.ifengidc.com/getUserHdhLogs?user=wangjf8&num=1000'
        res = requests.get(url = url)
        # print(res.status_code)
        # print(type(self.add_viewdata))
        # a = json.loads(self.add_viewdata)
        # print(res.json()[0]['brand'])
        a = res.json()[0]['logs']
        # b = json.dumps(a)
        # print(type(b))
        # print(a[1]['logs'][1]['opa'])
        # print(a)
        # assert 1 ==1
        return a

    def iot_assertion(self):
        # self.id_pinfo = ''
        for i in range(len(self.las)):
            # print('开始遍历las')
            # print(self.a)
            # print('标'+self.las[i][1])
            for j in range(len(self.a)):
                # print('地址'+self.a[j]['opa'])
                opa = self.las[i][1]
                record = self.las[i][2]
                aopa = self.a[j]['opa']
                arecord = self.a[j]['record']
                # print ('执行'+opa+':'+record)
                # print('a'+aopa+':'+arecord)
                try:
                    assert (opa) == (aopa)
                    assert (record) in (arecord)
                    print(opa+'匹配成功')
                    # self.id_pinfo = self.a[j]['record'][:22]

                    self.las[i][5] = self.a[j]['record']
                    
                    
                except Exception as e:
                    continue
        print(self.las)
#     def assertion_pegeinfo(self):
#         a = self.get_json()
#         for i in range(len(a)):
#             try:
#                 assert ('hdh_pageinfo') == (a[i]['opa'])
#                 assert ('article') in (a[i]['record'])
#                 print('pageinfo匹配成功')
#                 print(a[i]['opa']+':'+a[i]['record'][:22])  # pinfo
#             except Exception as e:
#                 print('不是pageinfo')

#     def assertion_pege(self):
#         a = self.get_json()
#         for i in range(len(a)):
#             try:
#                 assert ('hdh_page') == (a[i]['opa'])
#                 print('page匹配成功')
#                 print(a[i]['opa']+':'+a[i]['record'][:22])  # pinfo
#             except Exception as e:
#                 print('不是pege')

if __name__ == "__main__":
        unittest.main()
# if __name__ == '__main__':
    # statistics = Statistics()
    # bb = statistics.dowm()
    # aa = statistics.iot_assertion()

#     statistics.assertion()
